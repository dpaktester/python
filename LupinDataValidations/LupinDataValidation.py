import pyodbc
import os
import openpyxl

cnxnto36Db = pyodbc.connect(
    'DRIVER={SQL Server Native Client 10.0};SERVER=10.0.0.36;PORT=1433;DATABASE=M1B_LupRx;UID=m1bqc;PWD=m1bqc@0813')

cursorOn36Db = cnxnto36Db.cursor()

global file,wb,dataSheet
list_row = []
os.getcwd()
os.chdir("C:\\Users\\dmahapatra\\Desktop\\New folder")
wb = openpyxl.load_workbook("LupinQV.xlsx")
dataSheet = wb["DRUG_MATTY"]
PrecriberMattySheet=wb["Precriber_Matty"]
##summarySheet =  wb["Summary"]ReportOutput_Myrbetriq

    ##print("get all channels")
LupinQV_Drugs = []
TrxMatty = []
NrxMatty=[]
LupinQV_DrugMatty=[]

QVDrugMattyDic = {}
DBDrugMattyDic={}

##DrugMattyTRxDic.update({key: value})

"""
Compare The QV and Lite Drug Matty in Our Sheet
"""
def CheckQVandLiteMatty():
    count=0
    QVTRxMatty=[]
    LiteTRxMatty=[]
    QVNRxMatty=[]
    LiteNRxMatty=[]
    TRxDrugName=[]
    NRxDrugName=[]

    for i in range(5, dataSheet.max_row + 1):

        if str(dataSheet.cell(row=i, column=3).value)== str(dataSheet.cell(row=i, column=6).value):
            """print("")"""

        else:
            count =1
            TRxDrugName.append(str(dataSheet.cell(row=i, column=2).value))
            QVTRxMatty.append(str(round(dataSheet.cell(row=i, column=3).value,2)))
            LiteTRxMatty.append(str(round(dataSheet.cell(row=i, column=6).value,2)))

            #print("Failed for TRX")
            ###print("Valueof TRX is *********")

        if str(dataSheet.cell(row=i, column=9).value)== str(dataSheet.cell(row=i, column=12).value):

            """print("")"""
        else:
            count = 1
            NRxDrugName.append(str(dataSheet.cell(row=i, column=2).value))
            QVNRxMatty.append(str(round(dataSheet.cell(row=i, column=9).value,2)))
            LiteNRxMatty.append(str(round(dataSheet.cell(row=i, column=12).value,2)))
            ##print("Failed for NRX")
            ###print("Valueof NRX is *********")

    if (count==0):
        print("Test Case Passed for all Values")
        return 0
    else:
        print("Test Case Failed for Below Records")
        print("DrugName TRx  ",TRxDrugName.__str__())
        print(" TRx",QVTRxMatty)
        print("LiQVTRxMattyteTRxMatty TRx ",LiteTRxMatty)
        print("  ")
        print("DrugName NRx ", NRxDrugName)
        print("QVNRxMatty NRx ",QVNRxMatty)
        print("LiteNRxMatty NRx",LiteNRxMatty)
        return 1

#CheckQVandLiteMatty()

"""
Get the Drug Matty from our Sheet(QV)
"""
def QVOutPutTRx():
    global file, wb, dataSheet, QVDrugMattyDic
    for i in range(5, dataSheet.max_row+1 ):

###Make Key Value Pair of Drug and Trx and Nrx Matty and Update it in the Dictionary
        DrugsinReport = str(dataSheet.cell(row=i,column=2).value)

        if not ((dataSheet.cell(row=i, column=9).value  =="MATTY") | (dataSheet.cell(row=i, column=9).value  ==None) ):
            TRXandNRxMattyinReport = str(round(dataSheet.cell(row=i,column=3).value,2))+","+str(round(dataSheet.cell(row=i, column=9).value,2))
        QVDrugMattyDic.update({DrugsinReport: TRXandNRxMattyinReport})

    QVDrugMattyDic.pop("Womens Health Care")
    QVDrugMattyDic.pop("Drug Name")
    QVDrugMattyDic.pop("None")
    ##print("QVDrugMattyDic",QVDrugMattyDic)


#CheckQVandLiteMatty()

"""
Getting the DRUG Matty from Processed DB(RAW DB)
"""

def LupinProcessedDBMatty():

    """
    Change the Date of Month before checking the file.
    """
    global file, wb, dataSheet, DBDrugMattyDic

    DBQueryforMatty =    """     SELECT DrugS, SUM([TRX Matty]) AS TRX_DRUG_Matty ,SUM([NRX Matty])AS NRX_DRUG_Matty  FROM ( 
 SELECT  TRX [TRX Matty],NRX [NRX Matty],  
 CASE 
 WHEN PROD_GRP_ID IN ('00000101','00000104') THEN 'Suprax'
 WHEN PROD_GRP_ID IN ('00000110') THEN 'cefdinir'
 WHEN PROD_GRP_ID IN ('00000124','00000125') THEN 'amoxicillin cap'
 WHEN PROD_GRP_ID IN ('00000126','00000127','00000135','00000138') THEN 'amoxicillin tab'
 WHEN PROD_GRP_ID IN ('00000604','00000605') THEN 'Cleocin Vaginal'
 WHEN PROD_GRP_ID IN ('00000606') THEN 'Cleocin Vaginal Ovules'
 WHEN PROD_GRP_ID IN ('00000626') THEN 'Clindesse'
 WHEN PROD_GRP_ID IN ('00000627','00000628','00000629','00000630') THEN 'Flagyl'
 WHEN PROD_GRP_ID IN ('00000624') THEN 'clindamycin phosphate vaginal cream'
 WHEN PROD_GRP_ID IN ('00000634') THEN 'Metrogel-Vaginal' 
 WHEN PROD_GRP_ID IN ('00000635','00000636','00000637','00000638') THEN 'metronidazole'
 WHEN PROD_GRP_ID IN ('00000643') THEN 'metronidazole vaginal'
 WHEN PROD_GRP_ID IN ('00000647','00000648') THEN 'Tindamax'
 WHEN PROD_GRP_ID IN ('00000649','00000650','00000651','00000652') THEN 'tinidazole'
 WHEN PROD_GRP_ID IN ('00000697') THEN 'Solosec'
 WHEN PROD_GRP_ID IN ('00000698') THEN 'Nuvessa' END AS DrugS 
 FROM DFXponentAll WHERE Convert(DATE,period_id) BETWEEN '2018-06-01' AND '2019-05-31' 
 AND PROD_GRP_ID IN ('00000101','00000104','00000110','00000124','00000125','00000126','00000127','00000135','00000138','00000604','00000605','00000606','00000626',
 '00000627','00000628','00000629','00000630','00000624','00000634','00000635','00000636','00000637','00000638','00000643','00000647','00000648','00000649',
 '00000650','00000651','00000652','00000697','00000698'))A
 GROUP BY DrugS ORDER BY DrugS
                     """
    DBOutput = cnxnto36Db.execute(DBQueryforMatty)
    DBMattyforAllDrugs = DBOutput.fetchall()
    ##print("DB Output",DBMattyforAllDrugs)
    DatainDB=list(DBMattyforAllDrugs)

###1st Value is Drugname,2nd value is TRx Matty and 3rd Value is NRx Matty
    for i in range(len(DBMattyforAllDrugs)):
        count = 0
        DBDrugs = ""
        DBMAtty = ""
        for j in range(len(DBMattyforAllDrugs[i])):

            if(count == 0):
                DBDrugs = DBMattyforAllDrugs[i][j]
            else:
                tmp = round(DBMattyforAllDrugs[i][j],2)
                DBMAtty = DBMAtty+""+str(tmp)+","

            #print(DBMattyforAllDrugs[i][j])
            count= count+1

        DBDrugMattyDic.update({DBDrugs:DBMAtty.rstrip(",")})
    ###print(DBDrugMattyDic)
    ##key=DrugS


"""
Compare The Drug Matty between RAW DB and QV(Our File)
"""

def Comparison():
    count=0

    print("QVDrugMattyDic.items()",QVDrugMattyDic.items())
    print("DBDrugMattyDic.items().items()",DBDrugMattyDic.items())
    diffmatty =QVDrugMattyDic.items() - (DBDrugMattyDic.items())
    print("diffmatty",diffmatty)
    if (QVDrugMattyDic.__eq__(DBDrugMattyDic)):
        ###print(QVDrugMattyDic.items)
        #count=count +1
        print("The Test Case is passed", "\n", "All The Drug Matty in QV is Matching Against RAW DB")
    else:
        print("The Test Case is Failed for Below Drugs","\n")
        #print('\n',diffmatty)
        for item in diffmatty:
            print(item)
            count = count +1

    if count == 0:
        return 0
    else:
        return 1

#QVOutPutTRx()
#LupinProcessedDBMatty()
#Comparison()



def LupinPrecsriberMatty_TRx_Suprax_CapTab_Market_Raw_DB():

    """
    Change the Date of Month before checking the file and ME no if necessary
    """

    global DatainDB_Suprx_TRx

    DBQueryforTrxMatty_Suprax  = """ SELECT Sum(TRX) [TrxMatty_Suprax] from DFXponentAll WHERE Convert(Date,period_id) BETWEEN '2018-05-01' AND '2019-04-30'    and ME_Number ='0417701041'
                                     AND Prod_grp_ID in ('00000101','00000104','00000110','00000124','00000125','00000126','00000127','00000135','00000138')
                                 """
    DBOutput_Suprax_TRx = cnxnto36Db.execute(DBQueryforTrxMatty_Suprax)
    DbMattyforGivenPrescriber_TRx = DBOutput_Suprax_TRx.fetchall()
    ##print("DB Output",DBMattyforAllDrugs)
    ##DatainDB_Suprx_TRx = (round(DbMattyforGivenPrescriber_TRx),2)
    DatainDB_Suprx_TRx=round(DbMattyforGivenPrescriber_TRx[0][0],2)
    ##print("TrxMatty_Suprax",round(DatainDB_Suprx_TRx,2))

LupinPrecsriberMatty_TRx_Suprax_CapTab_Market_Raw_DB()
def LupinPrecsriberMatty_NRx_Suprax_CapTab_Market_Raw_DB():

    """
    Change the Date of Month before checking the file and ME no if necessary
    """
    global DatainDB_Suprx_NRx

    DBQueryforNrxMatty_Suprax  =  """ SELECT Sum(NRX) from DFXponentAll WHERE Convert(Date,period_id) BETWEEN '2018-05-01' AND '2019-04-30'   and ME_Number ='0417701041'
                                      AND Prod_grp_ID in ('00000101','00000104','00000110','00000124','00000125','00000126','00000127','00000135','00000138')
                                  """
    DBOutput_Suprax_NRx = cnxnto36Db.execute(DBQueryforNrxMatty_Suprax)
    DbMattyforGivenPrescriber_NRx = DBOutput_Suprax_NRx.fetchall()
    ##print("DB Output",DBMattyforAllDrugs)
    DatainDB_Suprx_NRx = round(DbMattyforGivenPrescriber_NRx[0][0], 2)
    ##DatainDB_Suprx_NRx = list(DbMattyforGivenPrescriber_NRx)


def LupinPrecsriberMatty_TRx_Womens_HealthCare_Market_Raw_DB():

    """
    Change the Date of Month before checking the file and ME no if necessary
    """
    global DBOutput_Womens_TRx

    DBQueryforTRxMatty_Womens  =  """   SELECT SUM(TRX) FROM DFXponentAll  WHERE Convert(Date,period_id) BETWEEN '2018-05-01' AND '2019-04-30'  AND  ME_Number = '217938203' AND PROD_GRP_ID IN 
                                       ('0000604','00000605','00000606',	'00000626','00000627','00000628','00000629','00000630',
                                       '00000624','00000634','00000635','00000636','00000637','00000638','00000643','00000647','00000648',
                                       '00000649','00000650','00000651','00000652','00000697','00000698')

                                  """
    DBOutput_Womens_TRx = cnxnto36Db.execute(DBQueryforTRxMatty_Womens)
    DbMattyforGivenPrescriber_NRx = DBOutput_Womens_TRx.fetchall()
    ##print("DB Output",DBMattyforAllDrugs)
    ##DBOutput_Womens_TRx = list(DbMattyforGivenPrescriber_NRx)
    DBOutput_Womens_TRx = round(DbMattyforGivenPrescriber_NRx[0][0], 2)

def LupinPrecsriberMatty_NRx_Womens_HealthCare_Market_Raw_DB():

    """
    Change the Date of Month before checking the file and ME no if necessary
    """
    global DBOutput_Womens_NRx
    DBQueryforNRxMatty_Womens  =  """   SELECT SUM(TRX) FROM DFXponentAll  WHERE Convert(Date,period_id) BETWEEN '2018-05-01' AND '2019-04-30'  AND  ME_Number = '217938203' AND PROD_GRP_ID IN 
                                       ('0000604','00000605','00000606',	'00000626','00000627','00000628','00000629','00000630',
                                       '00000624','00000634','00000635','00000636','00000637','00000638','00000643','00000647','00000648',
                                       '00000649','00000650','00000651','00000652','00000697','00000698')

                                  """
    DBOutput_Womens_NRx = cnxnto36Db.execute(DBQueryforNRxMatty_Womens)
    DbMattyforGivenPrescriber_NRx = DBOutput_Womens_NRx.fetchall()
    ##print("DB Output",DBMattyforAllDrugs)
    ##DBOutput_Womens_NRx = list(DbMattyforGivenPrescriber_NRx)
    DBOutput_Womens_NRx = round(DbMattyforGivenPrescriber_NRx[0][0], 2)


def PrecriberMatty():
    count=0
    SupraxCapTabTRxMatty=[]
    SupraxCapTabNRxMatty=[]
    MarketName_suprax=[]
    MarketName_Womens=[]
    WomensTRxMatty=[]
    WomensNRxMatty=[]
    LupinPrecsriberMatty_TRx_Suprax_CapTab_Market_Raw_DB()
    LupinPrecsriberMatty_NRx_Suprax_CapTab_Market_Raw_DB()
    LupinPrecsriberMatty_TRx_Womens_HealthCare_Market_Raw_DB()
    LupinPrecsriberMatty_NRx_Womens_HealthCare_Market_Raw_DB()


    MarketName_suprax.append(str(PrecriberMattySheet.cell(row=5, column=2).value))
    SupraxCapTabTRxMatty.append(str(round(PrecriberMattySheet.cell(row=5, column=5).value,2)))
    SupraxCapTabNRxMatty.append(str(round(PrecriberMattySheet.cell(row=6, column=5).value,2)))

    MarketName_Womens.append(str(PrecriberMattySheet.cell(row=7, column=2).value))
    WomensTRxMatty.append(str(round(PrecriberMattySheet.cell(row=7 , column=5).value,2)))
    WomensNRxMatty.append(str(round(PrecriberMattySheet.cell(row=8, column=5).value,2)))


    ###diff_of_SupraxCapTabTRxMatty = SupraxCapTabTRxMatty.items() - DatainDB_Suprx_TRx.items()
    if(SupraxCapTabTRxMatty.__contains__(str(DatainDB_Suprx_TRx))):
        print(" ")
    else:
        print("Test Case Failed")
        print("MarketName", MarketName_suprax)
        print ("SupraxCapTabTRxMatty",SupraxCapTabTRxMatty)
        print ("DatainDB_Suprx_TRx",DatainDB_Suprx_TRx)
        count=count+1
    if(SupraxCapTabNRxMatty.__contains__(str(DatainDB_Suprx_NRx))):
        print(" ")
    else:
        print("Test Case Failed")
        print("MarketName", MarketName_suprax)
        print("SupraxCapTabNRxMatty", SupraxCapTabNRxMatty)
        print("DatainDB_Suprx_NRx", DatainDB_Suprx_NRx)
        count = count + 1
    if (WomensTRxMatty.__contains__(str(DBOutput_Womens_TRx))):
        print(" ")
    else:
        print("Failed")
        print("MarketName", MarketName_Womens)
        print("WomensTRxMatty", WomensTRxMatty)
        print("DBOutput_Womens_TRx", DBOutput_Womens_TRx)
        count = count + 1
    if (WomensNRxMatty.__contains__(str(DBOutput_Womens_NRx))):
        print(" ")
    else:
        print("Failed")
        print("MarketName", MarketName_Womens)
        print("WomensNRxMatty", WomensNRxMatty)
        print("DBOutput_Womens_NRx", DBOutput_Womens_NRx)
        print("Difference of Matty is :")
        count = count + 1
    #print("SupraxCapTabTRxMatty",SupraxCapTabTRxMatty)
    #print("SupraxCapTabTRxMatty",SupraxCapTabNRxMatty)
    #print("WomensTRxMatty", WomensTRxMatty)
    #print("WomensNRxMatty", WomensNRxMatty)

    return count


##PrecriberMatty()