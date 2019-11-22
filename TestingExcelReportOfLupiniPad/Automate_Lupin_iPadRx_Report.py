import pyodbc
import os
import openpyxl

cnxn = pyodbc.connect(
    'DRIVER={SQL Server Native Client 10.0};SERVER=10.0.0.15;PORT=1433;DATABASE=M1B_Master;UID=m1bqc;PWD=m1b0813')

cursor = cnxn.cursor()

global file,wb,dataSheet,summarySheet,channelName,Drug
dicDetailssuprax = {}
dicDetailsolosec = {}
totalMatty=0
list_row = []
list_row = []
os.getcwd()
os.chdir("C:\\Users\\dmahapatra\\Desktop\\New folder")
wb = openpyxl.load_workbook("LupiniPadDataValiadtions_10252019.xlsx")
dataSheet1 = wb["National Aggregate"]
dataSheet2= wb["Prescriber Plan Level"]
channelNotIncludedSuprax=["Cash","HIX","Mail Order","VA","Other Third Party"]
channelNotIncludedSolosec=["Cash","HIX","Mail Order","VA","Other Third Party","Medicare"]

def verify__NationalAggregate():
    global file, wb, dataSheet,dicDetailssuprax,dicDetailsolosec
    print("get All the Channel's Matty")
    channelList = []
    DrugList = []
    totalCount = 0
    for  i in range(3, dataSheet1.max_row+1 ):
        channelList.append(str(dataSheet1.cell(row=i,column=1).value))
        DrugList.append(str(dataSheet1.cell(row=i, column=2).value))
    channelSet = set(channelList)
    DrugSet = set(DrugList)
    ##print(channelSet)
    ##print(DrugSet)
    chList = list(channelSet)
    drList = list(DrugSet)
    for eachChannel in range(chList.__len__()):
        channelName = chList[eachChannel]
        totalMatty = 0
        count = 0
        for i in range(4, dataSheet2.max_row+1 ):
            channel=dataSheet2.cell(row=i,column=6).value
            DrugName = dataSheet2.cell(row=i, column=7).value

            if((DrugName == "Suprax") and (channel == channelName) ):
                Matty = dataSheet2.cell(row=i, column=8).value
                #print(Matty)
                totalMatty = totalMatty + float(Matty)
                dicDetailssuprax.update({"Suprax"+"|"+channelName:round(totalMatty, 2)})

    for eachChannel in range(chList.__len__()):
        channelName = chList[eachChannel]
        totalMatty = 0
        count = 0
        for i in range(4, dataSheet2.max_row + 1):
            channel = dataSheet2.cell(row=i, column=6).value
            DrugName = dataSheet2.cell(row=i, column=7).value

            if ((DrugName == "Solosec") and (channel == channelName)):
                Matty = dataSheet2.cell(row=i, column=8).value
                # print(Matty)
                totalMatty = totalMatty + float(Matty)
                dicDetailsolosec.update({"Solosec" + "|" + channelName: round(totalMatty, 2)})
    ##print(dicDetails)


    for key in dicDetailssuprax.keys():
        for v in range(4, dataSheet1.max_row + 1):
            channel1 = dataSheet1.cell(row=v, column=1).value
            drug1 = dataSheet1.cell(row=v, column=2).value
            key1 = drug1 + "|" + channel1
            value = dataSheet1.cell(row=v, column=3).value

            if ((key == key1) & ~(channelNotIncludedSuprax.__contains__(channel1))):

                dicDetailssupraxvalue = dicDetailssuprax.get(key)

                if (dicDetailssupraxvalue == round(value, 2)):
                    print("#############Passed###########")
                    print("Channel", channel1)
                    print(("drug", drug1))
                    print("dicMatyValue----", dicDetailssupraxvalue)
                    print("value----", value)

                else:
                    if(dicDetailssupraxvalue - round(value, 2) > 0.5):
                        print("*********************Failed**********************")
                        #print(round(value, 2))
                        print("Channel", channel1)
                        print(("drug", drug1))
                        print("dicMatyValue----", dicDetailssupraxvalue)
                        print("key", key)
                        print("value----", value)
                        totalCount = totalCount+1

    for key in dicDetailsolosec.keys():
        for v in range(3, dataSheet1.max_row + 1):
            channel1 = dataSheet1.cell(row=v, column=1).value
            drug1 = dataSheet1.cell(row=v, column=2).value
            key1 = drug1 + "|" + channel1
            value = dataSheet1.cell(row=v, column=3).value
            if((key ==key1 ) & ~(channelNotIncludedSolosec.__contains__(channel1))):

                dicMatyValuesolosec = dicDetailsolosec.get(key)

                if (dicMatyValuesolosec == round(value, 2)):
                        print("#############Passed###########")
                        print("Channel", channel1)
                        print(("drug",drug1))
                        print("dicMatyValue----", dicMatyValuesolosec)
                        print("value----", round(value,2))
                        drugmattyCombination=(channel1,drug1,dicMatyValuesolosec,value)
                        writeToSheet("PASSED", "Check Matty difference between Precsriber tab and National Aggreage Tab", "",
                                     "Matty should match between both the tabs")

                else:
                    if (dicMatyValuesolosec - round(value, 2) > 0.5):
                        print("*********************Failed**********************")
                        print("*********************Failed**********************")
                        # print(round(value, 2))
                        print("Channel", channel1)
                        print(("drug", drug1))
                        print("dicMatyValue----", dicMatyValuesolosec)
                        print("key", key)
                        print("value----", round(value,2))
                        totalCount = totalCount + 1
                        drugmattyCombination = (channel1, drug1, dicMatyValuesolosec, value)
                        writeToSheet("PASSED",
                                     "Check Matty difference between Precsriber tab and National Aggreage Tab", "Matty is not matching for below records: ",drugmattyCombination,
                                     "Matty should match between both the tabs")


    return totalCount



#print(verify__NationalAggregate())


def DifferenceofMattyinRawDbandIpadDB():
    countofIssuesinComparisonofMatty = 0

    ReportQuery =    """  SELECT PM.PrescriberId,PM.FirstName,PM.LastName,PD.EntityID,PD.EntityName,PD.Channel,PD.DrugId,SUM(RX1+RX2+RX3+RX4+RX5+RX6+RX7+RX8+RX9+RX10+RX11+RX12)Matty  into #ReportResult2
                          FROM [B1-124].LupiniPadRx.DBO.PrescriberMaster PM 
                          JOIN [B1-124].LupiniPadRx.DBO.PrescriberDetails PD ON PM.PrescriberID=PD.PrescriberID
                          WHERE PD.TheraID IN(773,774,782) AND PD.DrugID 
                          ----IN(11539 /*Antara (Single Source) Drug*/,
                          IN (12870 /*Suprax Drug*/ ,12870 /*Solosec Drug*/ ,13140 /*Solosec*/)
                          GROUP BY PM.PrescriberId,PM.FirstName,PM.LastName,PD.EntityID,PD.EntityName,PD.Channel,PD.DrugId
                     """

    ReportOutput = cursor.execute(ReportQuery)

    RawDBQuery= """ SELECT A.SRA2 DEA_NO,A.ME_NO PRSC_ID,A.L_NM L_NM, A.F_NM F_NM,C.ENTITY_ID,C.CHANNEL_ID, B.THERA_ID,B.DRUG_BRD_ID,F.DRUG_BRAND_NM,
                    A.PRSC_ZIP ZIP,A.PRSC_ADDR ADDRESS,A.PRSC_CITY CITY,A.PRSC_STATE STATE,A.SRA3 SPECIALTY,'' TITLE,
                    --CASE WHEN A.SRA2=Z.PRSC_ID  THEN Z.FLAG ELSE NULL END TARGET_INDICATOR,
                    SUM(TRX1) RX1,SUM(TRX2) RX2,SUM(TRX3) RX3,SUM(TRX4) RX4,SUM(TRX5) RX5,SUM(TRX6) RX6,
                    SUM(TRX7) RX7,SUM(TRX8) RX8,SUM(TRX9) RX9,SUM(TRX10) RX10,SUM(TRX11) RX11,SUM(TRX12) RX12,
                    SUM(TRX1)+SUM(TRX2)+SUM(TRX3)+SUM(TRX4)+SUM(TRX5)+SUM(TRX6)+SUM(TRX7)+SUM(TRX8)+
                    SUM(TRX9)+SUM(TRX10)+SUM(TRX11)+SUM(TRX12) MATTY_RX  INTO #RawDataTotal
                    FROM [B1-123].M1B_LUPRX.DBO.EXT_RAW_DATA A 
                    LEFT OUTER JOIN (SELECT ZIP,MSA_ID,CBSA_DIV_ID FROM [B1-123].DI_MASTER.dbo.MHC_ZIP_CBSA) D ON A.PRSC_ZIP=D.ZIP
                    INNER JOIN [B1-123].DI_MASTER.dbo.MHC_EXT_DRUG_REL B 
                    ON A.PROD_ID=B.EXT_DRUG_ID  AND A.REPORT_NO=B.EXT_THERA_ID --included on 30 Nov 2017
                    AND B.CLIENT_ID = 2 AND B.STATUS='A' --AND B.THERA_ID=782 --PD is included on 7th march 2017
                    --LEFT OUTER JOIN [B1-123].M1B_LUPRX.DBO.MHC_PRSC_TARGET Z ON A.SRA2=Z.PRSC_ID AND Z.THERA_ID=B.THERA_ID
                    ---LEFT OUTER JOIN [B1-123].M1B_LUPRX.DBO.DF_PDRP_LIST M ON A.SRA2=M.IMS_NUM AND M.THERA_ID=B.THERA_ID --for other clients me_no was linked --THERA_ID CONDITION INCLUDED
                    LEFT JOIN [b1-124].M1B_MASTER.dbo.MHC_DRUG_BRAND F ON F.DRUG_BRAND_ID=B.DRUG_BRD_ID
                    INNER JOIN [B1-123].M1B_LUPRX.DBO.MHC_TERR_ZIP_RETL Y ON A.PRSC_ZIP=Y.ZIP AND Y.THERA_ID=B.THERA_ID,
                    (SELECT DISTINCT PAYER_PLAN_ID, X.ENTITY_ID, X.CHANNEL_ID
                    FROM [B1-123].DI_MASTER.dbo.TMP_ENTITY_IMS_XREF X WHERE X.CHANNEL_ID NOT IN (274,283,284))C
                    WHERE A.PAYER_PLAN_CD = C.PAYER_PLAN_ID --AND M.IMS_NUM IS NULL
                    AND (A.L_NM IS NOT NULL OR A.F_NM IS NOT NULL)  AND RTRIM(Ltrim(A.ME_NO))<>'0'
                    --AND DRUG_BRD_ID=12941
                    GROUP BY A.SRA2,A.ME_NO,A.L_NM, A.F_NM,C.ENTITY_ID,C.CHANNEL_ID, B.THERA_ID,B.DRUG_BRD_ID,F.DRUG_BRAND_NM,
                    A.PRSC_ZIP,A.PRSC_ADDR,A.PRSC_CITY,A.PRSC_STATE,A.SRA3--,(CASE WHEN A.SRA2=Z.PRSC_ID  THEN Z.FLAG ELSE NULL END)
                """

    RawDBOutPut = cursor.execute(RawDBQuery)

    DiffernceBetweenRawDBAndReportDb="""    SELECT DISTINCT A.PRSC_ID,B.EntityID,A.RawChannel,B.EntityName,B.Channel iPadChannel,
                                            A.Drug,A.[RAWDB_MATTY],B.Matty AS [iPADB_MATTY],[RAWDB_MATTY]-B.Matty AS Diff
                                            FROM (SELECT ENTITY_ID,PRSC_ID,RawChannel,DRUG_BRD_ID,DRUG,SUM(MATTY_RX) [RAWDB_MATTY]
                                            FROM (SELECT A.Entity_ID,A.PRSC_ID,A.DRUG_BRD_ID,
                                            CASE 
                                            WHEN A.CHANNEL_ID IN ('43005') THEN 'Tricare'
                                            WHEN A.CHANNEL_ID IN ('272','281') THEN 'Medicaid'
                                            WHEN A.CHANNEL_ID IN ('273','43007') THEN 'Commercial'
                                            WHEN A.CHANNEL_ID IN ('278','279','280') THEN 'Medicare' 
                                            WHEN A.CHANNEL_ID ='275' THEN 'PBM' END AS RawChannel,
                                            CASE 
                                            WHEN A.DRUG_BRD_ID IN ('12870') THEN 'Suprax'
								            WHEN A.DRUG_BRD_ID IN ('13140') THEN 'Solosec'
                                            -- WHEN A.DRUG_BRD_ID IN ('11539') THEN 'Antara' 
								            END AS DRUG,
                                            A.MATTY_RX
                                            FROM #RawDataTotal A                          
                                            WHERE A.DRUG_BRD_ID IN (11539,12870) AND A.CHANNEL_ID IN ('43005','272','281','273','278','279','280') 
                                            --and A.PRSC_ID='0160298023'
                                            )X
                                            GROUP BY ENTITY_ID,PRSC_ID,RawChannel,DRUG_BRD_ID,DRUG)A
                                            INNER JOIN #ReportResult2 B 
                                            ON A.PRSC_ID=B.PrescriberId AND A.DRUG_BRD_ID=B.DrugId AND A.RawChannel=B.Channel 
                                            AND A.Entity_ID=B.EntityID
                                            WHERE [RAWDB_MATTY]<>B.Matty

                                    """

    DiffernceBetweenRawDBAndReportDbOutput=cursor.execute(DiffernceBetweenRawDBAndReportDb)
    DifferenceofBothiPadandRawDb = DiffernceBetweenRawDBAndReportDbOutput.fetchall()

    #print("Difference between iPadDB and RawDB = ",DifferenceofBothiPadandRawDb)
    count=0
    if(DifferenceofBothiPadandRawDb.__len__()==0):
        writeToSheet("PASSED", "Check Matty difference between iPad DB and RAW DB Tab", "","There should not be any difference of matty between iPad DB and RAW DB")
        print("Prescriber Matty is Matching with RAW DB")

    else:
        writeToSheet("FAILED", "Check Matty difference between iPad DB and RAW DB Tab", "There are differences between ipad DB and Raw DB",
                     "There should not be any difference of matty between iPad DB and RAW DB")
        print("Failed")


def Verify_Channel_in_NationalAggregate():
    channelSet__in__File_list= ['Commercial','Medicare','Medicaid','TRICARE','Cash','HIX','Mail Order','Other Third Party','PBM','VA']
    channelSet__in__File_in_Nationaonal_Aggregateset=set(channelSet__in__File_list)
    ##print("Compare All Channel against Report")
    channelList = []
    for  i in range(4, dataSheet1.max_row+1 ):
        channelList.append(str(dataSheet1.cell(row=i,column=1).value))
    channelSet = set(channelList)
    difference_of_channel=set(channelSet__in__File_in_Nationaonal_Aggregateset)^set(channelSet)
    if difference_of_channel==set():
     count=0
    else:
        count=1
    if(count==0):
        writeToSheet("PASSED", "Check Channel difference between Natitional Aggregate Tab and Requirement Document", "",
                     "The Channel should match as per the requirement ")
    else:
        writeToSheet("FAILED", "Check Channel difference between Natitional Aggregate Tab and Requirement Document", "The Channel is not matching",
                     "The Channel should match as per the requirement ")


def Verify_Drugs_in_NationalAggregate():
    drugS__in__File_list= ['Suprax','Solosec','amoxicillin cap','amoxicillin tab','cefdinir','Cleocin Vaginal','Cleocin Vaginal Ovules',
                                'clindamycin phosphate vaginal cream','Clindesse','Flagyl','Metrogel-Vaginal','metronidazole','metronidazole vaginal','Nuvessa',
                                'Tindamax',	'tinidazole']
    drugS__in__File_set=set(drugS__in__File_list)
    ##print("Compare All Drugs against Report")
    DrugList = []
    for  i in range(4, dataSheet1.max_row+1 ):
        DrugList.append(str(dataSheet1.cell(row=i, column=2).value))
    DrugSet = set(DrugList)
    difference_of_drugs=set(drugS__in__File_list)^set(DrugSet)

    if difference_of_drugs==set():
     count=0
    else:

        count=1
    if(count==0):
        print("Drugs are matching agaisnt the Report")
        writeToSheet("PASSED", "Check The Drugs against Requiremnt for National Aggregate Tab", "",
                     "Drugs should match agaisnt the Report for Prescriber_Plan_Level tab")
    else:
        print("TestCase Failed ","\n","Difference of Drugs are", difference_of_drugs)
        writeToSheet("FAILED", "Check The Drugs against Requiremnt for National Aggregate Tab",
                     "Differtence of Drugs are ::", difference_of_drugs,
                     "Drugs should match agaisnt the Report for National Aggregate tab")


def Verify_Channel_in__Prescriber_Plan_Level_Sheet():
    channelSet__in__File_list = ['Commercial', 'Medicare', 'Medicaid', 'TRICARE','PBM']
    channelSet__in__File_set = set(channelSet__in__File_list)
    channelList = []
    DrugList = []
    for i in range(4, dataSheet2.max_row + 1):
        channelList.append(str(dataSheet2.cell(row=i, column=6).value))
        DrugList.append(str(dataSheet2.cell(row=i, column=7).value))
    DrugSet = set(DrugList)
    channelSet = set(channelList)
    difference_of_channel = set(channelSet__in__File_set) ^ set(channelSet)
    if difference_of_channel == set():
        count = 0
    else:
        count = 1
    if (count == 0):
        print("TestCase Passed","\n","Channels are matching agaisnt the Report for National Aggregate Tab")
        writeToSheet("PASSED", "Check The Channels against Requiremnt for National Aggregate Tab", "",
                     "Drugs should match agaisnt the Report for Prescriber_Plan_Level tab")
    else:
        print("TestCase Failed ", "\n", "Difference of Channels are", difference_of_channel)
        writeToSheet("FAILED", "Check The Channels against Requiremnt for National Aggregate Tab", "Differtence of Channels are ::"+difference_of_channel.__str__(),
                     "Channels should match agaisnt the Report for National Aggregate tab")


def Verify_Drugs_in_Prescriber_Plan_Level_Sheet():
    drugS__in__File_list_in_Prescriber_Plan_Level= ['Suprax','Solosec']
    drugS__in__File_set_in_Prescriber_Plan_Level=set(drugS__in__File_list_in_Prescriber_Plan_Level)
    ##print("Compare All Drugs against Report")
    DrugList = []
    for  i in range(4, dataSheet2.max_row+1 ):
        DrugList.append(str(dataSheet2.cell(row=i, column=7).value))
    DrugSet = set(DrugList)
    difference_of_drugs=set(drugS__in__File_set_in_Prescriber_Plan_Level)^set(DrugSet)

    if difference_of_drugs==set():
     count=0
    else:

        count=1
    if(count==0):
        print("TestCase Passed ","\n","Drugs are matching agaisnt the Report for Prescriber_Plan_Level tab")
        writeToSheet("PASSED", "Check The Drugs against Requiremnt", "",
                     "Drugs should match agaisnt the Report for Prescriber_Plan_Level tab")
    else:
        print("TestCase Failed ","\n","Difference of Drugs are for Prescriber_Plan_Level tab", difference_of_drugs)
        writeToSheet("FAILED", "Check The Drugs against Requiremnt", "Drugs are not matching Againt Requiremnt,difference_of_drugs :: ",difference_of_drugs,
                     "Drugs should match agaisnt the Report for Prescriber_Plan_Level tab")


def Verify_RawDB_Matty_Agaisnt_IpadDB_Matty_in_Prescriber_Plan_Level_Sheet():
    count=0
    diffence_in_prescribers=""
    for  i in range(4, dataSheet2.max_row+1 ):
        RAWDBmatty1= str(dataSheet2.cell(row=i, column=8).value)
        iPadDBmatty2 =str(dataSheet2.cell(row=i, column=9).value)
        if not( RAWDBmatty1 == iPadDBmatty2):
            diffence_in_prescribers=diffence_in_prescribers+","+str(dataSheet2.cell(row=i, column=1).value)+"|"+str(dataSheet2.cell(row=i, column=6).value)\
                            +"|"+str(dataSheet2.cell(row=i, column=7).value)+"|"+str(dataSheet2.cell(row=i, column=5).value)
            count=1

    if(count==0):
            print("TestCase Passed ","\n","RAwDB Matty is matching agaisnt the iPAD DB in Prescriber_Plan_Level tab")
            writeToSheet("PASSED", "Check RawDB Matty against iPadDB", "",
                         "The RawDB Matty should Match Against iPad Matty")
    else:
            print("TestCase Failed for  " + diffence_in_prescribers,"  Prescriebr ")
            writeToSheet("FAILED", "Check RawDB Matty against iPad DB", "The RawDB Matty is not Matching against iPadDB  Matty",
                         "The RawDB Matty shouold Match Against iPad Matty")

def Verify_BlankEntity__in_Prescriber_Plan_Level_Sheet():
    count=0
    blankEntitiesinfo = ""
    for  i in range(4, dataSheet2.max_row+1 ):
        entityName= str(dataSheet2.cell(row=i, column=5).value)
        #print("entityName",entityName)
        if ( entityName == 'None'):
            blankEntitiesinfo = blankEntitiesinfo+","+str(dataSheet2.cell(row=i, column=1).value)+"|"+str(dataSheet2.cell(row=i, column=6).value)\
                            +"|"+str(dataSheet2.cell(row=i, column=7).value)+"|"+str(dataSheet2.cell(row=i, column=5).value)
            count=1

    if(count==0):
            print("There is no Blank Rows in Entity Name Field of Prescriber Plan Level Sheet ")
            writeToSheet("PASSED", "Check Blank Field in Entity Name Field of Prescriber Plan Level Sheet", "",
                         "There should not be Blank Rows in Prescriber ME Number Field of Prescriber Plan Level Sheet")
    else:
            print("There is a Blank Rows in Entity Name Field of PEntity Name Field for Prescriber ID ",blankEntitiesinfo)
            writeToSheet("FAILED", "Check Blank Field in Entity Name Field of Prescriber Plan Level Sheet", "There is Blank Field in Entity Field",
                         "There should not be Blank Rows in Prescriber ME Number Field of Prescriber Plan Level Sheet")

def Verify_Blank_Precriber_in_Prescriber_Plan_Level_Sheet():
    count=0
    blankPrescriberMeNos = ""
    for  i in range(4, dataSheet2.max_row+1 ):
        PrescriberMeNos= str(dataSheet2.cell(row=i, column=1).value)
        #print("PrescriberMeNos",PrescriberMeNos)
        if ( PrescriberMeNos == 'None'):
            blankPrescriberMeNos = blankPrescriberMeNos+"  "+str(dataSheet2.cell(row=i, column=1).value)+"|"+str(dataSheet2.cell(row=i, column=2).value)+"|"+str(dataSheet2.cell(row=i, column=3).value)\
                              +"|"+str(dataSheet2.cell(row=i, column=6).value)+"|"+str(dataSheet2.cell(row=i, column=7).value)+"|"+\
                              str(dataSheet2.cell(row=i, column=5).value+"\n")
            count=1

    if(count==0):
            print("There is no Blank Rows in Prescriber ME Number Field of Prescriber Plan Level Sheet ")
            writeToSheet("PASSED", "Check Blank Field in Prescriber ME Number Field of Prescriber Plan Level Sheet", "",
                         "There should not be Blank Rows in Prescriber ME Number Field of Prescriber Plan Level Sheet")
    else:

            writeToSheet("FAILED", "Check Blank Field in Prescriber ME Number Field of Prescriber Plan Level Sheet", "There is Blank Filed in Prescribe ME Field",
                         "There should not be Blank Rows in Prescriber ME Number Field of Prescriber Plan Level Sheet")

def verify__Antara_in_NationalAggregate_Sheet():
    count = 0
    HavingRecordswithAntara= ""
    Static_Information_NationalAggregate = str(dataSheet1.cell(row=1, column=1).value)
    # print("PrescriberMeNos",PrescriberMeNos)
    if (Static_Information_NationalAggregate.__contains__("Antara (single source)")):
           # HavingRecordswithAntara=str(dataSheet2.cell(row=1, column=1).value)
            count = 1
            #print("The Sheet should not have the Drug Antara (single source) in the file",HavingRecordswithAntara)

    if (count == 0):
        print("There is no Antara Drug in NationalAggregate Sheet ")
        writeToSheet("PASSED", "Check Antara Drug in NationalAggregate Tab", "", "Antara Drug should not be there in National Aggregate Tab")
    else:
        print("There is Antara Drug in the Static field of Prescriber Plan Level Sheet", "\n", Static_Information_NationalAggregate, "\n")
        writeToSheet("FAILED", "Check Antara Drug in NationalAggregate Tab", "Antara Drug is there in National Aggregate Tab", "Antara Drug should not be there in National Aggregate Tab")

def verify__Antara_in_Prescriber_Plan_Level_sheet():
    count = 0
    Static_Information_Prescriber_Plan_Level_sheet = str(dataSheet2.cell(row=1, column=1).value)
    # print("PrescriberMeNos",PrescriberMeNos)
    if (Static_Information_Prescriber_Plan_Level_sheet.__contains__("Antara (single source)")):
           # HavingRecordswithAntara=str(dataSheet2.cell(row=1, column=1).value)
            count = 1
            #print("The Sheet should not have the Drug Antara (single source) in the file",HavingRecordswithAntara)

    if (count == 0):
        print("There is no Antara Drug in Prescriber Plan Level Sheet ")
        writeToSheet("PASSED", "Check Antara Drug in Prescriber Plan Level Sheet", "","Antara Drug should not be there in Prescriber Plan Level Sheet")
    else:
        writeToSheet("FAILED", "Check Antara Drug in Prescriber Plan Level Sheet", "Antara Drug found in Prescriber Plan Level Sheet ","Antara Drug should not be there in Prescriber Plan Level Sheet")


###Create Excel Code####
wb1 = openpyxl.Workbook(write_only=True)
def createWorkBook():
        global  wb1
        print("create")
        os.getcwd()
        print("create work sheet")
        wb1.save('TestReportOfLupiniPadRx.xlsx')
        wb1 = openpyxl.load_workbook("TestReportOfLupiniPadRx.xlsx")


def createSheet():
        print("sheet")
        wb1.create_sheet("TestAutomationReport")
        wb1.save("TestReportOfLupiniPadRx.xlsx")

def writeHeaderToSheet():


    sheet = wb1["TestAutomationReport"]
    print("rows",sheet.max_row)


    sheet.cell(row = 1, column=1).value  = "TEST CASE STATUS"
    sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row = 1, column=2).value = "TEST CASE NAME"
    sheet.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)

    sheet.cell(row = 1, column=3).value = "TEST FAIL DESCRIPTION"
    sheet.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row = 1, column=4).value = "EXPECTED RESULT"
    sheet.cell(row=1, column=4).font = openpyxl.styles.Font(bold=True)

    wb1.save("TestReportOfLupiniPadRx.xlsx")

def writeToSheet(testStatus,testCaseName,description,exceptedResult):

    sheet =wb1["TestAutomationReport"]

    len = sheet.max_row+1
    #print("Length",len,testStatus,testCaseName,description)

    sheet.cell(row=len, column=1).value = testStatus
    #print(sheet.cell(row=len, column=1).value)

    sheet.cell(row=len, column=2).value = testCaseName
    #print(sheet.cell(row=len, column=2).value)

    #print(sheet.cell(row=len, column=3).value)
    sheet.cell(row=len, column=3).value = description
    sheet.cell(row=len, column=4).value = exceptedResult
    wb1.save("TestReportOfLupiniPadRx.xlsx")

def closeWorkBook():
        global wb
        wb1.save("TestReportOfLupiniPadRx.xlsx")
        wb1.close()




createWorkBook()
createSheet()
writeHeaderToSheet()

verify__NationalAggregate()
Verify_Channel_in_NationalAggregate()
Verify_Drugs_in_NationalAggregate()
Verify_Channel_in__Prescriber_Plan_Level_Sheet()
Verify_Drugs_in_Prescriber_Plan_Level_Sheet()
Verify_RawDB_Matty_Agaisnt_IpadDB_Matty_in_Prescriber_Plan_Level_Sheet()
Verify_BlankEntity__in_Prescriber_Plan_Level_Sheet()
Verify_Blank_Precriber_in_Prescriber_Plan_Level_Sheet()
verify__Antara_in_NationalAggregate_Sheet()
verify__Antara_in_Prescriber_Plan_Level_sheet()
DifferenceofMattyinRawDbandIpadDB()

closeWorkBook()