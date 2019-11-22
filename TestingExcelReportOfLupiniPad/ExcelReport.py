import openpyxl
import os
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import BLUE


class ExcelReports:


    def createWorkBook(self,ExcelName):

        print("create")
        os.getcwd()
        print("create work sheet")
        wb = openpyxl.Workbook(write_only=True)
        wb.save('TOB_DataBaseScripts.xlsx')
        wb = openpyxl.load_workbook("TOB_DataBaseScripts.xlsx")
        return wb


    def createSheet(self,sheetName,wb):
        print("sheet",sheetName)
        wb.create_sheet(sheetName)
        wb.save("TOB_DataBaseScripts.xlsx")



    def writeHeaderToSheet(self,sheetName,wb):


        sheet = wb[sheetName]
        print("rows",sheet.max_row)


        sheet.cell(row = 1, column=1).value  = "TEST CASE STATUS"
        sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row = 1, column=2).value = "TEST CASE NAME"
        sheet.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)

        sheet.cell(row = 1, column=3).value = "TEST FAIL DESCRIPTION"
        sheet.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row = 1, column=4).value = "EXPECTED RESULT"
        sheet.cell(row=1, column=4).font = openpyxl.styles.Font(bold=True)

        wb.save(self.fileName)

    def writeToSheet(self,sheetName,wb):
        fileName = self.fileName
        sheet =wb[sheetName]

        len = sheet.max_row+1
        #print("Length",len,testStatus,testCaseName,description)

        sheet.cell(row=len, column=1).value = self.testStatus
        #print(sheet.cell(row=len, column=1).value)
        if(self.testStatus == "Failed"):

          sheet.cell(row=len, column=1).font = openpyxl.styles.Font(color=RED)
        else:

          sheet.cell(row=len, column=1).font = openpyxl.styles.Font(color=BLUE)

        sheet.cell(row=len, column=2).value = self.testCaseName
        #print(sheet.cell(row=len, column=2).value)

        #print(sheet.cell(row=len, column=3).value)
        sheet.cell(row=len, column=3).value = self.description
        sheet.cell(row=len, column=4).value = self.exceptedResult
        wb.save(fileName)

    def closeWorkBook(self):
        global wb
        wb.save("Acella_pb_report.xlsx")
        wb.close()

