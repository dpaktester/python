import openpyxl
import os
import _pydecimal
from decimal import Decimal
import pyodbc
import re
ReportOutput_Myrbetriq=[]
ReportOutput_Myrbetriq_XL=[]
ReportOutput_Vesicare=[]
ReportOutput_Vesicare_XL=[]
valuet=[]
cnxnto15Db = pyodbc.connect('DRIVER={SQL Server Native Client 10.0};SERVER=10.0.0.15;PORT=1433;DATABASE=AstellasPB;UID=m1bqc;PWD=m1b0813')
cursor = cnxnto15Db.cursor()
global file,wb,dataSheet,summarySheet
list_row = []
os.getcwd()
os.chdir("C:\\Users\\dmahapatra\\Desktop\\November 2018 PB Supporting Documents")
wb = openpyxl.load_workbook("Myrbetriq-Vesicare_Access by Channel June 2019.xlsx")
dataSheet = wb["Details"]
##summarySheet =  wb["Summary"]ReportOutput_Myrbetriq



def verify_total_Myrbetriq():
    global file, wb, dataSheet,ReportOutput_Myrbetriq_XL
    ##print("get all channels")
    channelList = []
    statusList = []
    for  i in range(4, dataSheet.max_row+1 ):
        channelList.append(str(dataSheet.cell(row=i,column=12).value))
        statusList.append(str(dataSheet.cell(row=i, column=19).value))
    channelSet = set(channelList)
    statusSet = set(statusList)
    ##print(channelSet)
    ## print(statusSet)
    chList = list(channelSet)
    stList = list(statusSet)
    for eachChannel in range(chList.__len__()):
        channelName = chList[eachChannel]
        for eachStatus in range(stList.__len__()):
            status = stList[eachStatus]
            totalRxLivesMyrbetriq = 0
            totalRxLivesVesicare=0
            #print(status)
            for i in range(1, dataSheet.max_row+1 ):
                channel=dataSheet.cell(row=i,column=12).value
                drugstatus = dataSheet.cell(row=i,column=19).value
                DrugName = dataSheet.cell(row=i, column=18).value
                if(drugstatus == status and DrugName == "Myrbetriq" and channel == channelName ):
                      #print("totalRxLives", totalRxLives, "channel", channelName, "STatus", status)
                      rxValue = dataSheet.cell(row=i, column=16).value
                      #print("rxValue",rxValue)
                      if(rxValue == None):
                          rxValue = 0

                      totalRxLivesMyrbetriq =totalRxLivesMyrbetriq+ int(rxValue)

                if (drugstatus == status and DrugName == "VESIcare" and channel == channelName):
                   # print("totalRxLivesVesicare", totalRxLivesVesicare, "channel", channelName, "STatus", status,"Drugname",DrugName)
                    rxValue = dataSheet.cell(row=i, column=16).value
                   # print("rxValue",rxValue)
                    if (rxValue == None):
                        rxValue = 0

                    totalRxLivesVesicare = totalRxLivesVesicare + int(rxValue)

                    ## print("totalRxLives",totalRxLives,"channel",channelName,"STatus",status)



           ## print("totalRxLives",totalRxLives,"channel",channelName,"STatus",status)

            ReportOutput_Myrbetriq_XL.append(channelName+"|"+str(totalRxLivesMyrbetriq)+"|"+status)
            ##print("ReportOutput_Myrbetriq_XL",ReportOutput_Myrbetriq_XL)

            ReportOutput_Vesicare_XL.append(channelName + "|" + str(totalRxLivesVesicare) + "|" + status)
            ##print("ReportOutput_Vesicare_XL", ReportOutput_Vesicare_XL)

##verify_total_Myrbetriq()

def PB_For_Myrbetriq():

    PB_Table_Query_Myrbetriq =   """   SELECT EF.CustomSubChannel,sum(FormularyRxLives)[Sum OF FormularyRx Lives],DF.DrugStatus FROM  EntityFormulary EF 
                        JOIN DrugFormularyStatus DF ON EF.EntityID=DF.EntityID AND EF.FormularyID=DF.FormularyID 
                        WHERE DF.DrugName='MyrBetriq'-- AND DF.DrugStatus='PB'
                        GROUP BY EF.CustomSubChannel,DF.DrugStatus
                        ORDER BY EF.customsubChannel,DF.DrugStatus
                    """
    Table_Query_Myrbetriq_Output1 = cursor.execute(PB_Table_Query_Myrbetriq)
    Table_Query_Myrbetriq_Output= Table_Query_Myrbetriq_Output1.fetchall()
    for each in range(list(Table_Query_Myrbetriq_Output).__len__()):
        ReportOutput_Myrbetriq.append(Table_Query_Myrbetriq_Output[each][0] +"|"+ str(Table_Query_Myrbetriq_Output[each][1]) +"|"+ Table_Query_Myrbetriq_Output[each][2] )
        ##print(ReportOutput_Myrbetriq)


def PB_For_Vesicare():

    PB_Table_Query_Vesicare =   """   SELECT EF.CustomSubChannel,sum(FormularyRxLives)[Sum OF FormularyRx Lives],DF.DrugStatus FROM  EntityFormulary EF 
                        JOIN DrugFormularyStatus DF ON EF.EntityID=DF.EntityID AND EF.FormularyID=DF.FormularyID 
                        WHERE DF.DrugName='Vesicare'-- AND DF.DrugStatus='PB'
                        GROUP BY EF.CustomSubChannel,DF.DrugStatus
                        ORDER BY EF.customsubChannel,DF.DrugStatus
                    """
    Table_Query_Vesicare_Output1 = cursor.execute(PB_Table_Query_Vesicare)
    Table_Query_Vesicare_Output= Table_Query_Vesicare_Output1.fetchall()
    for each in range(list(Table_Query_Vesicare_Output).__len__()):
        ReportOutput_Vesicare.append(Table_Query_Vesicare_Output[each][0] +"|"+ str(Table_Query_Vesicare_Output[each][1]) +"|"+ Table_Query_Vesicare_Output[each][2] )
        ##print(ReportOutput_Vesicare)



##PB_For_Myrbetriq()

def Compare_Myrbetriq_Report():
    global ReportOutput_Myrbetriq
    PB_For_Myrbetriq()
    verify_total_Myrbetriq()
    count = 0
    for i in range(ReportOutput_Myrbetriq.__len__()):
        item = ReportOutput_Myrbetriq[i]
        if not (ReportOutput_Myrbetriq_XL.__contains__(item)):

            count = count + 1
            print("Test Case is failed")
            ##print(item)
            print("ReportOutput_Myrbetriq",ReportOutput_Myrbetriq)
            print("ReportOutput_Myrbetriq_XL",ReportOutput_Myrbetriq_XL)



    if(count == 0):
        print(" Access Report Data is matching Against Astellas Payerbackbone Data for Myrbetriq Drug ")

    ##print(set(ReportOutput_Myrbetriq).difference(set(ReportOutput_Myrbetriq_XL)))
    return count

def Compare_Vesicare_Report():
    global ReportOutput_Vesicare
    PB_For_Vesicare()
    verify_total_Myrbetriq()
    count = 0
    for i in range(ReportOutput_Vesicare.__len__()):
        item = ReportOutput_Vesicare[i]
        if not(ReportOutput_Vesicare_XL.__contains__(item)):
            print(" ")
            count = count + 1
            print("Test Case is failed")
            ##print(item)
            print("ReportOutput_Vesicare",ReportOutput_Vesicare)
            print("ReportOutput_Vesicare_XL",ReportOutput_Vesicare_XL)


    if (count == 0):
        print(" Access Report Data is matching Against Astellas Payerbackbone Data for VESIcare Drug ")

    ##print(set(ReportOutput_Myrbetriq).difference(set(ReportOutput_Myrbetriq_XL)))
    return count


#Compare_Myrbetriq_Report()
#Compare_Vesicare_Report()

def Check_0_ID__in_National_Entity_ID():
    global file, wb, dataSheet
    count=0
    for i in range(3, dataSheet.max_row + 1):
        Entity__ID=(str(dataSheet.cell(row=i, column=1).value))
        if(Entity__ID=="0"):
            Parent__entity__name = str(dataSheet.cell(row=i, column=2).value)
            Astellas__Parent__entity__name = str(dataSheet.cell(row=i, column=3).value)
            if ((Astellas__Parent__entity__name == "None") & (Parent__entity__name == "None")):
                if (count==0):
                    print("Test Case Passed",'\n',"Parent entity name is blank for National ID is 0",'\n',"Astellas Parent entity name is blank for National ID 0")
                    return 0
                else:
                    print("Test Case Failed")
                    print("Astellas__Parent__entity__name",Astellas__Parent__entity__name)
                    print("Parent__entity__name",Parent__entity__name)
                    return 1

Check_0_ID__in_National_Entity_ID()



def Check_alphabets_in_NumericFields():
    global file, wb, dataSheet
    count=0
    numericerror=""
    for i in range(4, dataSheet.max_row + 1):
            NationalEntityID=(str(dataSheet.cell(row=i, column=1).value))
            RegionalEntityID=(str(dataSheet.cell(row=i, column=4).value))
            EntityID=(str(dataSheet.cell(row=i, column=7).value))
            Planlives=(str(dataSheet.cell(row=i, column=13).value))
            PlanRxlives=(str(dataSheet.cell(row=i, column=14).value))
            FormularyRxlives=(str(dataSheet.cell(row=i, column=16).value))
            DrugTier =(str(dataSheet.cell(row=i, column=20).value))


            if ((re.findall("[a-zA-Z]", NationalEntityID).__len__()!=0) | (re.findall("[a-zA-Z]", RegionalEntityID).__len__()!=0) |
                    (re.findall("[a-zA-Z]", EntityID).__len__()!=0)| (re.findall("[a-zA-Z]", Planlives).__len__()!=0 and Planlives!="None") |
                    (re.findall("[a-zA-Z]", PlanRxlives).__len__()!=0 and PlanRxlives !="None") | (re.findall("[a-zA-Z]", FormularyRxlives).__len__()!=0 and FormularyRxlives != "None")
                    | (re.findall("[a-zA-Z]", DrugTier).__len__()!=0 and DrugTier != "None")):


                numericerror=numericerror+"|"+"NationalEntityID"+"|"+NationalEntityID+"|"+"RegionalEntityID"+"|"+RegionalEntityID+"|"+"EntityID"+"|"+EntityID+"|"+"Planlives"\
                             +"|"+Planlives+"|"+"PlanRxlives"+"|"+PlanRxlives+"|"+"FormularyRxlives"+"|"+FormularyRxlives+"|"+"DrugTier"+"|"+DrugTier+"\n"+"\n"
                #print("Issues are:----",numericerror)
                count=1
    if (count == 0):
        print("Test Case Passed", '\n', "There is no alphabetical charecters in all the numeric fields")
        return 0
    else:
        print("Test Case Failed","\n","There is issue with these data  ","\n",numericerror)
        print("Count", count)
        return 1




Check_alphabets_in_NumericFields()
"""
              if (count==0):
                    print("Test Case Passed",'\n',"There is no alphabetical charecters in all the numeric fieleds")
                    return 0
                else:
                    print("Test Case Failed")
                    print("NationalEntityID",NationalEntityID)
                    print("RegionalEntityID",RegionalEntityID)
                    print("EntityID", EntityID)
                    print("Planlives", Planlives)
                    print("PlanRxlives", PlanRxlives)
                    print("FormularyRxlives", FormularyRxlives)
                    print("DrugTier", DrugTier)


                    return 1

"""


def Check_AstellasChannel():
    global file, wb, dataSheet
    count=0
    AstellasChannel_list=[]
    AstellasChannel_set=[]
    defaultChannel_list=['HIX','Commercial','Managed Medicaid','Medicare','State Medicaid','TRICARE','VA']
    defaultChannel_set=set(defaultChannel_list)
    for i in range(4, dataSheet.max_row + 1):
        AstellasChannel=(str(dataSheet.cell(row=i, column=11).value))
        #AstellasSubChannel = (str(dataSheet.cell(row=i, column=12).value))
        AstellasChannel_list.append(AstellasChannel)
        #print("AstellasChannel", AstellasChannel)
        #print("AstellasSubChannel", AstellasSubChannel)
    AstellasChannel_set=set(AstellasChannel_list)
    difference_of_channel = AstellasChannel_set.difference(defaultChannel_set)
    #print(AstellasChannel_set)

    if difference_of_channel == set():
        count = 0
    else:
        count = 1
    if (count == 0):
        print("TestCase Passed","\n","Channels are matching Against Requirement")
        return 0
    else:
        print("TestCase Failed ", "\n", "Difference of Channels are", difference_of_channel)
        return 1

##Check_AstellasChannel()

def Check_Astellas_Sub_Channel():
    global file, wb, dataSheet
    count = 0

    for i in range(4, dataSheet.max_row + 1):

        AstellasChannel = dataSheet.cell(row=i, column=11).value
        AstellasSubChannel = dataSheet.cell(row=i, column=12).value
        if ((AstellasChannel == 'Commercial') or (AstellasChannel == 'HIX') or (
                AstellasChannel == 'Managed Medicaid') or (AstellasChannel == 'Medicare') or (
                AstellasChannel == 'State Medicaid')
                or (AstellasChannel == 'TRICARE') or (AstellasChannel == 'VA')):
            # or (AstellasChannel=='Managed Medicaid')):
            if  ((AstellasChannel == 'Commercial' and AstellasSubChannel == 'Commercial' or AstellasSubChannel == 'Employer')
                or (AstellasChannel == 'HIX' and AstellasSubChannel == 'HIX')
                or (AstellasChannel == 'Managed Medicaid' and AstellasSubChannel == 'Managed Medicaid')
                or (AstellasChannel == 'Medicare' and AstellasSubChannel == 'MA-PD' or AstellasSubChannel == 'PDP')
                or (AstellasChannel == 'State Medicaid' and AstellasSubChannel == 'State Medicaid')
                or (AstellasChannel == 'TRICARE' and AstellasSubChannel == 'TRICARE')
                or (AstellasChannel == 'VA' and AstellasSubChannel == 'VA')):

                # print(i,"AstellasChannel",AstellasChannel,"AstellasSubChannel",AstellasSubChannel)
                # print("Passed")
                count = 0
            else:
                #print("failed")
                print("Row Number","|",i,"|","AstellasChannel","|",AstellasChannel,"|","AstellasSubChannel","|",AstellasSubChannel)
                count = count + 1
                #print("count", count)
    if (count == 0):
        print("Test Case Passed ","\n","The Sub Channels are matching as per the Channels")
        return 0
    else:
        print("\n","Test Case Failed")
        return 1



def Check_Blank_ID__in_Regional_Entity_ID():
    global file, wb, dataSheet
    count=0
    error = ""
    for i in range(4, dataSheet.max_row + 1):
        Regional_Entity_ID = str(dataSheet.cell(row=i, column=4).value)
        National_Entity_ID = str(dataSheet.cell(row=i, column=1).value)

        if ((Regional_Entity_ID=='None') | (re.findall("[a-zA-Z]", Regional_Entity_ID).__len__()!=0)):

            error = str(error+str(National_Entity_ID)+"|"+str(Regional_Entity_ID))+"\n"
            #print(i,"Regional_Entity_ID", Regional_Entity_ID)
            count=1
        #else:
            #count=0
    if(count==0):
            print("Test Case Passed",'\n',"There is no Blank Regoinal EntityID")
            return 0
    else:
            print("Test Case Failed",'\n',"National ID and Regional ID is ::","\n",error)
            return 1

#Check_Blank_ID__in_Regional_Entity_ID()