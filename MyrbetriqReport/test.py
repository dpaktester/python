import openpyxl
import os
import _pydecimal
from decimal import Decimal
import pyodbc

ReportOutput_Myrbetriq=[]
ReportOutput_Myrbetriq_XL=[]
ReportOutput_Vesicare=[]
ReportOutput_Vesicare_XL=[]
valuet=[]
cnxnto15Db = pyodbc.connect('DRIVER={SQL Server Native Client 10.0};SERVER=10.0.0.15;PORT=1433;DATABASE=AstellasPB;UID=m1bqc;PWD=m1b0813')
cursor = cnxnto15Db.cursor()
global file,wb,dataSheet,summarySheet
list_row = []
os.getcwd()
os.chdir("C:\\Users\\dmahapatra\\Desktop\\November 2018 PB Supporting Documents")
wb = openpyxl.load_workbook("Myrbetriq-Vesicare_Access.xlsx")
dataSheet = wb["Details"]
##summarySheet =  wb["Summary"]ReportOutput_Myrbetriq

def Check_Astellas_Sub_Channel():
    global file, wb, dataSheet
    count=0

    for i in range(4, dataSheet.max_row + 1):

        AstellasChannel=dataSheet.cell(row=i, column=11).value
        AstellasSubChannel= dataSheet.cell(row=i, column=12).value
        if ((AstellasChannel=='Commercial') or (AstellasChannel=='HIX') or (AstellasChannel=='Managed Medicaid') or(AstellasChannel=='Medicare') or (AstellasChannel=='State Medicaid')
            or (AstellasChannel=='TRICARE') or (AstellasChannel=='VA')):
            #or (AstellasChannel=='Managed Medicaid')):
            if((AstellasChannel=='Commercial' and AstellasSubChannel =='Commercial' or AstellasSubChannel=='Employer') or (AstellasChannel== 'HIX' and AstellasSubChannel=='HIX')
                                  or (AstellasChannel=='Managed Medicaid' and AstellasSubChannel=='Managed Medicaid') or (AstellasChannel=='Medicare' and AstellasSubChannel=='MA-PD' or AstellasSubChannel=='PDP')
                                  or (AstellasChannel=='State Medicaid' and AstellasSubChannel=='State Medicaid')
                                  or (AstellasChannel=='TRICARE' and AstellasSubChannel=='TRICARE')
                                  or (AstellasChannel=='VA' and AstellasSubChannel=='VA')):


                #print(i,"AstellasChannel",AstellasChannel,"AstellasSubChannel",AstellasSubChannel)
                #print("Passed")
                count=0
            else:
                print("failed")
                print(i, "Else AstellasChannel", AstellasChannel, "Else AstellasSubChannel", AstellasSubChannel)
                count=count+1
                print("count",count)
    if (count==0):
                return 0
                ##print("Passed")
    else:
                print("failed")
                return  1

print(Check_Astellas_Sub_Channel())
'''
   if (count == 0):
        print("TestCase Passed", "\n", "Sub Channels are matching Against Requirement for Commercial Channel")
        print()
        return 0
    else:
        print("TestCase failed", "\n", "Sub Channels are Not matching Against Requirement for Channel", "\n",
              "Differnce of Sub Channels for are", "AstellasChannel", AstellasChannel, "AstellasSubChannel",
              AstellasSubChannel)
        return 1
'''


