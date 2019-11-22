import openpyxl
import os
import _pydecimal
from decimal import Decimal
import pyodbc

ReportOutput_Myrbetriq=[]
ReportOutput_Vesicare=[]
valuet=[]
cnxnto15Db = pyodbc.connect('DRIVER={SQL Server Native Client 10.0};SERVER=10.0.0.15;PORT=1433;DATABASE=AstellasPB;UID=m1bqc;PWD=m1b0813')
cursor = cnxnto15Db.cursor()
global file,wb,dataSheet,summarySheet
list_row = []
os.getcwd()
os.chdir("C:\\Users\\dmahapatra\\Desktop\\November 2018 PB Supporting Documents\\November 2018 PB Supporting Documents")
wb = openpyxl.load_workbook("Myrbetriq-Vesicare_Access by Channel November 2018.xlsx")
dataSheet = wb["Details"]
##summarySheet =  wb["Summary"]ReportOutput_Myrbetriq


def verify_total_Myrbetriq():
    global file, wb, dataSheet
    print(dataSheet.max_row)
    for i in range(1, dataSheet.max_row+1 ):
        valueS=dataSheet.cell(row=i,column=1).value


verify_total_Myrbetriq()

def PB_For_Myrbetriq():

    PB_Table_Query_Myrbetriq =   """   SELECT EF.CustomSubChannel,sum(FormularyRxLives)[Sum OF FormularyRx Lives],DF.DrugStatus,DF.CustomDrugStatus FROM  EntityFormulary EF 
                        JOIN DrugFormularyStatus DF ON EF.EntityID=DF.EntityID AND EF.FormularyID=DF.FormularyID 
                        WHERE DF.DrugName='MyrBetriq'-- AND DF.DrugStatus='PB'
                        GROUP BY DF.CustomDrugStatus,EF.CustomSubChannel,DF.DrugStatus
                        ORDER BY EF.customsubChannel,DF.DrugStatus
                    """
    Table_Query_Myrbetriq_Output1 = cursor.execute(PB_Table_Query_Myrbetriq)
    Table_Query_Myrbetriq_Output= Table_Query_Myrbetriq_Output1.fetchall()
    for each in range(list(Table_Query_Myrbetriq_Output).__len__()):
        ReportOutput_Myrbetriq.append(Table_Query_Myrbetriq_Output[each][0] +"|"+ str(Table_Query_Myrbetriq_Output[each][1]) +"|"+ Table_Query_Myrbetriq_Output[each][2] +"|"+ Table_Query_Myrbetriq_Output[each][3])
        ##print(ReportOutput_Myrbetriq)


##PB_For_Myrbetriq()


def Compare_Myrbetriq_Report():
    global ReportOutput_Myrbetriq
    PB_For_Myrbetriq()
    verify_total_Myrbetriq()
    count = 0
    for i in range(ReportOutput_Myrbetriq.__len__()):
        item = ReportOutput_Myrbetriq[i]
        if valuet.__contains__(item):
            print("")
        else:

            count = count + 1


    return count

